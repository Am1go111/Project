import telebot
import time
from datetime import datetime
from openpyxl import load_workbook
import sqlite3
from telebot import types
token="6313526890:AAGI7lKDwb4IKzyAeNLktLsuLLhNNa1f1cQ"
bot=telebot.TeleBot(token)
name = None
surname = None
podrazdel = None
fn = 'DB.xlsx'
wb = load_workbook(fn)
ws = wb['data']
now = datetime.now()
current_time = now.strftime("%H:%M")


def checkData():
  now = datetime.now()
  date = datetime(now.year, now.month, 4)
  if date.isoweekday() == 6 or date.isoweekday() == 7:
    return 0
  else:
    return 1


@bot.message_handler(commands=['start'])
def start(message):
  f = False
  tr = False
  while f == False:
    if checkData() == 0 and tr == False:
      tr = True
      bot.send_message(message.chat.id,"Привет! Сегодня у меня выходной, желаю провести время с удовольствием и наполниться ресурсом для предстоящей рабочей недели, обязательно напишу тебе в понедельник. Хороших выходных!")
      time.sleep(3600)
    elif checkData() == 1:
      f = True
  conn = sqlite3.connect('BD.sql')
  cur = conn.cursor()
  cur.execute('CREATE TABLE IF NOT EXISTS users (id int auto_increment primary key, name varchar(50),'
              ' surname varchar(50), pod varchar(50), dolsh varchar(50))')
  conn.commit()
  cur.close()
  conn.close()
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while now.hour < 9 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Добрый день, коллега!  Добро пожаловать в Cosmos Hotel Group (CHG)!  Мы верим, что великий путь сотрудника начинается с адаптации. Меня зовут COSMI 🚀 моя миссия - сделать адаптацию каждого сотрудника на новом рабочем месте комфортной и плавно погрузить  в корпоративную жизнь через важную информацию, блиц задания и интересные факты. На протяжении двух недель я буду на связи, ко мне можно обращаться с вопросами, а я постараюсь помочь, для этого нужно нажать кнопку «Задать вопрос».", reply_markup= markup)
  bot.send_message(message.chat.id,"Здорово, что этот день проведем в едином потоке, отправляю лучи тепла и отличного настроения! :)")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Дорогой коллега, начнем со знакомства! Ниже предложены несколько полей для заполнения информации о себе")
  bot.send_message(message.chat.id,"Напишите ваше ФИО")
  bot.register_next_step_handler(message,user_name)

def user_name(message):
  global name
  name = message.text.strip()
  bot.send_message(message.chat.id, "Напишите свою корпоративную почта")
  bot.register_next_step_handler(message,user_surname)

def user_surname(message):
  global surname
  surname = message.text.strip()
  bot.send_message(message.chat.id, "Напишите свое подразделение")
  bot.register_next_step_handler(message, user_pod)

def user_pod(message):
  global podrazdel
  podrazdel = message.text.strip()
  bot.send_message(message.chat.id, "Напишите свою должность")
  bot.register_next_step_handler(message, user_dol)

def user_dol(message):
  dolsh = message.text.strip()
  conn = sqlite3.connect('BD.sql')
  cur = conn.cursor()
  cur.execute('INSERT INTO users (name, surname, pod, dolsh) VALUES ("%s","%s","%s","%s")'%(name,surname,podrazdel,dolsh))
  cur.execute('SELECT * FROM  users')
  users = cur.fetchall()
  ws[f'A{str(len(users))}'] = message.chat.id
  ws[f'B{str(len(users))}'] = name
  ws[f'C{str(len(users))}'] = surname
  ws[f'D{str(len(users))}'] = podrazdel
  ws[f'E{str(len(users))}'] = dolsh
  wb.save(fn)
  conn.commit()
  cur.close()
  conn.close()
  bot.send_message(message.chat.id,"Каждый день я буду присылать новые и интересные задания")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"<b>Задание на сегодня:</b> HR команда CHG уже отправила на корпоративную почту ПРИВЕТСТВЕННОЕ ПИСЬМО, предлагаю сегодня:", parse_mode='html')
  bot.send_message(message.chat.id,"1) Изучить ПРИВЕТСТВЕННОЕ ПИСЬМО")
  bot.send_message(message.chat.id,"2) Принять приглашение-регистрацию на корпоративный портал")
  bot.send_message(message.chat.id,"3) Зайти в корпоративный портал")
  bot.send_message(message.chat.id, "Если возникла трудность на данном этапе, можно нажать кнопку «Задать вопрос» или написать на почту hr@cosmosgroup.ru")
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id, "Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, первый рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!")
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Привет! Команда CHG желает доброго и чудесного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Cosmos Hotel Group (CHG) – приобрела у норвежской Wenaas Hotel Russia AS компаний-владельцев десяти отелей, включая Park Inn by Radisson, в четырех российских городах. По итогам сделки CHG стала владельцем 6 отелей в Петербурге, 2 в Москве, 1 в Мурманске, 1 - в Екатеринбурге. Сделка стала крупнейшей за всю историю гостиничного рынка России.")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Задание на сегодня: Оформить корпоративную подпись в почте. Перейти в папку U:\Общие документы - Правила жизни офиса, выбрать файл  «Правила жизни офиса_ПОДПИСЬ», ознакомиться с ним,  в Outlook создать свою корпоративную подпись.")
  time.sleep(10)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да"))
  markup.add(types.KeyboardButton("Нет"))
  bot.send_message(message.chat.id,"Получилось ли справиться с заданием?",reply_markup=markup)
  bot.register_next_step_handler(message,day2o)
def day2o(message):
  otvet = message.text.strip()
  if otvet.lower() == "да":
    k = 1
    bot.send_message(message.chat.id,"Бгодарим за открытость- это бесценно для всех коллег!")
    while message.chat.id != (ws[f'A{str(k)}'].value):
      k+=1
    ws[f'F{str(k)}'] = otvet
    wb.save(fn)
  elif otvet.lower() == "нет":
    k = 1
    bot.send_message(message.chat.id, "Возникли трудности? Можно выбрать кнопку «Задать вопрос», я с удовольствием постараюсь помочь")
    while message.chat.id != (ws[f'A{str(k)}'].value):
      k += 1
    ws[f'G{str(k)}'] = otvet
    wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while (now.hour < 16  and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!",reply_markup=markup)
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Привет! Команда CHG желает доброго и чудесного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "Корпоративная культура в CHG играет очень важную роль, поэтому отправляю в чат правила нашего взаимодействия, благодаря которым адаптация пройдет быстрее:")
  bot.send_message(message.chat.id, "Работа в команде, с командой и за команду")
  bot.send_message(message.chat.id, "Соблюдение деловой этики")
  bot.send_message(message.chat.id, "Уважительное отношение к коллегам")
  bot.send_message(message.chat.id, "Соблюдение режима рабочего времени")
  bot.send_message(message.chat.id, "Соблюдение дресс-кода ")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "Задание на сегодня: Зайти на корпоративный портал, кликнуть на свой профиль (в верхнем правом углу), добавить фотографию и заполнить контактную информацию о себе. ")
  time.sleep(10)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да"))
  markup.add(types.KeyboardButton("Нет"))
  bot.send_message(message.chat.id, "Получилось ли справиться с заданием?", reply_markup=markup)
  bot.register_next_step_handler(message, otv)
def otv(message):
  otvet = message.text.strip()
  if otvet.lower() == "да":
    k = 1
    bot.send_message(message.chat.id,"Бгодарим за открытость- это бесценно для всех коллег!")
    while message.chat.id != (ws[f'A{str(k)}'].value):
      k+=1
    ws[f'H{str(k)}'] = otvet
    wb.save(fn)
  elif otvet.lower() == "нет":
    k = 1
    bot.send_message(message.chat.id, "Возникли трудности? Можно выбрать кнопку «Задать вопрос», я с удовольствием постараюсь помочь")
    while message.chat.id != (ws[f'A{str(k)}'].value):
      k += 1
    ws[f'I{str(k)}'] = otvet
    wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!",reply_markup= markup)
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Доброго дня! Больше всего сил у нас появляется тогда, когда нам нравится то, что мы делаем. Желаю, чтобы сегодняшний день принес море позитива!")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Задание на сегодня: Дорогой, коллега! Команда CHG идет в ногу со временем и с 28.06.2023 перешла на электронный документооборот в сфере трудовых отношений (КЭДО) на базе информационной системы - СБИС (Sabymy). Вам уже была направлена ссылка и инструкция для регистрации в личном кабинете СБИС.")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да"))
  markup.add(types.KeyboardButton("Нет"))
  bot.send_message(message.chat.id, "Получилось ли зарегистрироваться?", reply_markup=markup)
  bot.register_next_step_handler(message,blank)
def blank(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'J{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!",reply_markup= markup)
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 17) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"COSMI на связи! Желаю прекрасного утра и мощного заряда позитивом.")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Ознакомиться с порядком оформления командировок. Для этого необходимо зайти на «диск U-общие документы - 11. КЭДО СБИС_КОГ». Чуть позже я отправлю квиз по этой теме.")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Квиз: В ответе необходимо указать правильную последовательность процессов организации Командировки (где 1- это первый шаг, а 8 –финальный шаг)?  ")
  bot.send_message(message.chat.id, "1.	Войти в личный кабинет Sabymy СБИС, в раздел «Документы».")
  bot.send_message(message.chat.id, "2.	Выбрать вариант командировки (общественный транспорт, личное авто).")
  bot.send_message(message.chat.id, "3.	Заполнить все необходимые поля и отправить документ на выполнение.")
  bot.send_message(message.chat.id, "4.	Подписать Служебное задание и Приказ на командировку в личном кабинете Sabymy СБИС.")
  bot.send_message(message.chat.id, "5.	Заказать билеты и гостиницу через Cosmos Travel путем оформления Письма-заявки в Cosmos Travel.")
  bot.send_message(message.chat.id, "6.	Сохранить все оригиналы отчетных документов, квитанций и чеков по командировке.")
  bot.send_message(message.chat.id, "7.	Заполнить отчет по командировке в личном кабинете Sabymy СБИС.")
  bot.send_message(message.chat.id, "8.	Вложить оригиналы документов по командировке в конверт и передать бухгалтеру, Марине Маркиной (через лоток для отправки в гостиницу «ГК Космос» на 2 этаже).")
  bot.send_message(message.chat.id, "Ответ напишите одним сообщением")
  bot.register_next_step_handler(message,wifi)
def wifi(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'K{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  bot.send_message(message.chat.id,"Ответ: расположены в верном порядке.",reply_markup= markup)
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!")
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Привет! Желаю потрясающего, продуктивного, наполненного и просто чудесного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Задание на сегодня: Изучить инструкцию по оформлению отпуска в личном кабинете Sabymy СБИС. Для этого необходимо перейти на «диск U-общие документы - 11. КЭДО СБИС_КОГ», изучить инструкцию, отметить в комментариях какие виды отпусков Вы можете оформить через личный кабинет Sabymy СБИС.")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"CHG за непрерывное развитие каждого сотрудника. Все актуальные новости по этой теме, а также из корпоративной жизни компании, можно найти в разделе «Новости» на корпоративном портале. Коллеги всегда рады лайкам и позитивным комментариям!")
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!")
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Привет! Желаем доброго и бодрого утра, крутого и интересного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "CHG аккумулирует на корпоративном портале не только крутые новости, но и структуру   -  Управляющей компании и Отелей. Рекомендую в разделе «Компания» подробно ознакомиться со структурой, подразделениями, сотрудниками. Поделюсь классной фишкой – если на корпоративном портале в верхней строке поиска завести ФИ работающего сотрудника, сразу же найдется карточка его личного контакта. Пробуем!")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "Задание на сегодня: Ознакомиться с Правилами жизни офиса. Для этого необходимо зайти на «диск U-общие документы» и изучить раздел «Правила жизни офиса». Ниже в строке ответа, нужно указать к какому гостевому Wi-fi можно предложить гостю подключиться при нахождении в офисе CHG. ")
  bot.register_next_step_handler(message, disk)
def disk(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'L{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!",reply_markup= markup)
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id, "Привет! Желаю замечательного и ясного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "Задание на сегодня: Зайти на сайт cosmosgroup.ru, перейти в раздел о нас и обратить особое внимание на вкладки «О компании», «Отели» и «Бренды». Завтра оправлю интересный квиз по темам, которые представлены на сайте компании.")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id, "В CHG запущен корпоративный телеграмм канал. Все самое актуальное – свежие новости, публикации в прессе, интересные события и акции отелей сети CHG.Присоединяйтесь! https://t.me/cosmoshotelgroup")
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!")
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)
  bot.send_message(message.chat.id,"COSMI на связи! Доброго утра, чудесного настроения и самочувствия!")
  while now.hour < 14 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Задание на сегодня: Пройти квиз и проверить свои знания о компании.")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("2003"))
  markup.add(types.KeyboardButton("2006"))
  markup.add(types.KeyboardButton("2018"))
  bot.send_message(message.chat.id,"1)Год основания компании:", reply_markup=markup)
  bot.register_next_step_handler(message,per)
def per(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'M{str(k)}'] = otvet
  wb.save(fn)
  bot.send_message(message.chat.id,"Ответ: 2003")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Новый Уренгой, Томск, Саратов"))
  markup.add(types.KeyboardButton("Мурманск, Грозный, Ижевск"))
  markup.add(types.KeyboardButton("Сегежа, Улан-Удэ, Когалым"))
  bot.send_message(message.chat.id, "2) В каких городах РФ НЕ представлены отели CHG:", reply_markup=markup)
  bot.register_next_step_handler(message, vtor)
def vtor(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'N{str(k)}'] = otvet
  wb.save(fn)
  bot.send_message(message.chat.id, "Ответ: Новый Уренгой, Томск, Саратов")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("В любом месте веселее вместе "))
  markup.add(types.KeyboardButton("Открой вселенную гостеприимства"))
  markup.add(types.KeyboardButton("Выше звезд только Космос"))
  bot.send_message(message.chat.id,"3) Верный слоган компании:", reply_markup=markup)
  bot.register_next_step_handler(message, tret)
def tret(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'O{str(k)}'] = otvet
  wb.save(fn)
  bot.send_message(message.chat.id, "Ответ: Открой вселенную гостеприимства")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Cosmos Collections, Cosmos Hotels, Cosmos Smart"))
  markup.add(types.KeyboardButton("Cosmos Awards, Cosmos Tourist, Cosmos Messenger"))
  markup.add(types.KeyboardButton("Cosmos Travel, Cosmos Selection, Cosmos Stay"))
  bot.send_message(message.chat.id, "4) Выбери НЕсуществующие бренды CHG:", reply_markup=markup)
  bot.register_next_step_handler(message, chetvert)
def chetvert(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'P{str(k)}'] = otvet
  wb.save(fn)
  bot.send_message(message.chat.id, "Ответ: Cosmos Awards, Cosmos Tourist, Cosmos Messenger")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Организация и проведение банкетов"))
  markup.add(types.KeyboardButton("Организация и обучение аниматоров"))
  markup.add(types.KeyboardButton("Организация въездного туризма в России"))
  bot.send_message(message.chat.id,"5) Cosmos Travel в CHG это",reply_markup=markup)
  bot.register_next_step_handler(message,finalquiz)
def finalquiz(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'Q{str(k)}'] = otvet
  wb.save(fn)
  bot.send_message(message.chat.id, "Ответ: Организация въездного туризма в России")
  bot.send_message(message.chat.id,"Поздравляю с прохождением квиза! Надеюсь, сайт компании погрузил во многие корпоративные аспекты и дополнил представление о CHG. Свою обратную связь по сайту можно отправить в наш чат. ")
  bot.send_message(message.chat.id,"Напиши одним сообщением,если нет пожеланий,то напиши 'все хорошо'")
  bot.register_next_step_handler(message,day3)
def day3(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'R{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  while (now.hour < 16 and now.minute < 29) or now.hour >= 18:
    time.sleep(600)
  bot.send_message(message.chat.id,"Дорогой, коллега! В нашей корпоративной культуре принято задавать вопросы своему руководителю, коллегам и HR команде. Надеюсь, рабочий день прошел продуктивно и интересно, на все вопросы нашлись ответы. Спасибо за прекрасный день в компании!",reply_markup=markup)
  time.sleep(7200)
  while (now.hour < 9 or now.hour >= 18) or checkData() == 0:
    time.sleep(3600)or checkData() == 0
  bot.send_message(message.chat.id,"COSMI на связи! Желаю доброго и бодрого утра, сочного и интересного дня!")
  while now.hour < 11 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Задание на сегодня: CHG подключена к Alpina digital. Это самая большая бизнес-библиотека на русском языке, в том числе в мобильном приложении. На Alpina Digital можно найти онлайн книги, базу знаний, крутые и актуальные лекции и вебинары.Необходимо перейти по ссылке, https://alpina-digital.app.link/6ZcgedjC1ub , зарегистрироваться на сайте, используя корпоративную почту.")
  while now.hour < 13 or now.hour >= 18:
    time.sleep(3600)
  bot.send_message(message.chat.id,"Дорогой, коллега! Мнение каждого нового сотрудника очень важно для COSMI!  Пожалуйста, поделись своим мнением по результатам первых двух недель адаптации в компании. Направляю блиц опрос. Ответы помогут сделать адаптацию в CHG еще более комфортной и приятной.")
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да! Все отлично, есть все инструменты для эффективной работы"))
  markup.add(types.KeyboardButton("Ну… рабочий стол есть, но всех необходимых инструментов пока нет, работа протекает не всегда неэффективно"))
  markup.add(types.KeyboardButton("Нет, мое рабочее место не подготовлено, работаю, как получается."))
  bot.send_message(message.chat.id, "1.	Мое рабочее место подготовлено надлежащим образом:", reply_markup=markup)
  bot.register_next_step_handler(message, lasto)
def lasto(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'S{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да, я знаю к чему иду, и что должен сделать на испытательном периоде"))
  markup.add(types.KeyboardButton("Я примерно понимаю свои задачи, четкого определения цели у меня не было"))
  markup.add(types.KeyboardButton("Нет, мы не обсуждали с руководителем, я сам определяю, что мне необходимо сделать."))
  bot.send_message(message.chat.id, "2.	Мне четко сформулировали мои основные задачи:", reply_markup=markup)
  bot.register_next_step_handler(message, lastt)
def lastt(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'T{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да"))
  markup.add(types.KeyboardButton("Частично"))
  markup.add(types.KeyboardButton("Нет"))
  bot.send_message(message.chat.id, "3.	Я ознакомлен с культурой компании", reply_markup=markup)
  bot.register_next_step_handler(message, lasttr)
def lasttr(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'U{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да, уже хорошо знаю своих коллег"))
  markup.add(types.KeyboardButton("Пока в процессе"))
  markup.add(types.KeyboardButton("Нет, лишние знакомства мне не нужны"))
  bot.send_message(message.chat.id, "4.	Я познакомился со своими коллегами:", reply_markup=markup)
  bot.register_next_step_handler(message, lastf)
def lastf(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'V{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  markup.add(types.KeyboardButton("Да, у меня есть все необходимые контакты"))
  markup.add(types.KeyboardButton("Я всегда все спрашиваю у руководителя"))
  markup.add(types.KeyboardButton("Нет, я не знаю к кому обращаться по вопросам."))
  bot.send_message(message.chat.id, "5.	Я знаю, к кому обращаться, если у меня возникают различные вопросы:", reply_markup=markup)
  bot.register_next_step_handler(message, lastlast)
def lastlast(message):
  otvet = message.text.strip()
  k = 1
  while message.chat.id != (ws[f'A{str(k)}'].value):
    k += 1
  ws[f'W{str(k)}'] = otvet
  wb.save(fn)
  markup = types.ReplyKeyboardMarkup()
  btn1 = types.KeyboardButton("Не работает принтер")
  btn2 = types.KeyboardButton("Не работает Wi-fi")
  markup.row(btn1, btn2)
  btn3 = types.KeyboardButton("Вопрос по документам")
  btn4 = types.KeyboardButton("Не могу установить программу ")
  markup.row(btn3, btn4)
  btn5 = types.KeyboardButton("Не работает VPN")
  btn6 = types.KeyboardButton("Контакты HR")
  markup.row(btn5, btn6)
  btn7 = types.KeyboardButton("Кажется, я заболел")
  btn8 = types.KeyboardButton("Меня игнорирует руководитель")
  markup.row(btn7, btn8)
  btn9 = types.KeyboardButton("Вопрос по комплаенсу")
  btn10 = types.KeyboardButton("Задать вопрос")
  markup.row(btn9, btn10)
  bot.send_message(message.chat.id,"Дорогой, коллега! Сегодня последний день адаптации с Cosmi. Надеюсь наше взаимодействие было комфортным и полезным. Благодарю за вовлеченность и отзывчивость. Желаю отличного старта в компании и реализации самых интересных проектов. Удачи!",reply_markup=markup)











@bot.message_handler(content_types='text')
def message_reply(message):
  if checkData() == 1:
    if message.text == "Не работает принтер":
      bot.send_message(message.chat.id, "Напиши на Helpdesk COG helpdeskcog@cosmosgroup.ru")
    elif message.text == "Не работает Wi-fi":
      bot.send_message(message.chat.id, "Напиши на Helpdesk COG helpdeskcog@cosmosgroup.ru")
    elif message.text == "Вопрос по документам":
      bot.send_message(message.chat.id, "Напиши на Helpdesk COG helpdeskcog@cosmosgroup.ru")
    elif message.text == "Не могу установить программу":
      bot.send_message(message.chat.id, "Напиши на Helpdesk COG helpdeskcog@cosmosgroup.ru")
    elif message.text == "Не работает VPN":
      bot.send_message(message.chat.id, "Напиши на Helpdesk COG helpdeskcog@cosmosgroup.ru")
    elif message.text == "Контакты HR":
      bot.send_message(message.chat.id, "Напиши на hr@cosmosgroup.ru")
    elif message.text == "Кажется, я заболел":
      bot.send_message(message.chat.id, "Напиши на hr@cosmosgroup.ru")
    elif message.text == "Меня игнорирует руководитель":
      bot.send_message(message.chat.id, "Напиши на hr@cosmosgroup.ru")
    elif message.text == "Вопрос по комплаенсу":
      bot.send_message(message.chat.id, "Напиши на hr@cosmosgroup.ru")
    elif message.text == "Задать вопрос":
      bot.send_message(message.chat.id, "Напишите ваше сообщение,а я отправлю его отвественному сотруднику")
      global help_user_id
      help_user_id = message.from_user.id
      markup = types.InlineKeyboardMarkup()
      button1 = types.InlineKeyboardButton(text='Отмена', callback_data='cancel')
      markup.add(button1)
      msg = bot.send_message(message.chat.id, 'Задайте вопрос боту.', reply_markup=markup)
      bot.register_next_step_handler(msg, helpbot)
  else:
        bot.send_message(message.chat.id, "Привет! Сегодня у меня выходной, желаю провести время с удовольствием и наполниться ресурсом для предстоящей рабочей недели, обязательно напишу тебе в понедельник. Хороших выходных")

def helpbot(msg):
  bot.forward_message(6755650947, msg.chat.id, msg.message_id)

def handle_text(m):
   if int(m.chat.id) == int(6755650947):
      bot.send_message(help_user_id, m.text)



bot.infinity_polling()