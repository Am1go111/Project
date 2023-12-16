import telebot
import sqlite3
from telebot import types
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
token="6877115430:AAGs2U8aZUUgB0FAyOIqlvWcoobztFw3gik"
bot=telebot.TeleBot(token)
fn = 'DB.xlsx'
wb = load_workbook(fn)
ws = wb['data']
name = None
age = None
city = None

@bot.message_handler(commands=['start'])

def start(message):
    """
    Запускает бота,создает Базу данных
    :param message: Имя
    :return: Имя пользователя
    """
    conn = sqlite3.connect('BD.sql')
    cur = conn.cursor()
    cur.execute('CREATE TABLE IF NOT EXISTS users (id int auto_increment primary key, name varchar(50),'
                ' age varchar(50), city varchar(50))')
    conn.commit()
    cur.close()
    conn.close()
    bot.send_message(message.chat.id,"Привет я бот,который будет присылть тебе гороскоп")
    bot.send_message(message.chat.id, "Напиши ,пожалуйста, как тебя зовут?")
    bot.register_next_step_handler(message,user_name)

def user_name(message):
    """
    Записывает имя пользователя для дальнейшего использования,узнает возраст пользователя
    :param message: Возраст
    :return:  Возраст пользователя
    """
    global name
    name = message.text.strip()
    bot.send_message(message.chat.id, f"А теперь подскажи мне свой возраст, {name}?")
    bot.register_next_step_handler(message,agee)

def agee(message):
    """
    Записывает возраст пользователя для дальнейшего использования,узнает из какого города пользоватеь
    :param message: Город
    :return: Место проживания пользователя
    """
    global age
    age = message.text.strip()
    bot.send_message(message.chat.id, f"А теперь подскажи мне,где ты живешь {name} в свои {age}?")
    bot.register_next_step_handler(message, cityy)

def cityy(message):
    """
    Переносит данные в таблицу Excell,создает кнопки для дальнейшей работы
    :param message: Знак зодиака
    :return: Знак зодиака пользователя
    """
    global city
    city = message.text.strip()
    conn = sqlite3.connect('BD.sql')
    cur = conn.cursor()
    cur.execute('INSERT INTO users (name, age, city) VALUES ("%s","%s","%s")' % (name, age, city))
    cur.execute('SELECT * FROM  users')
    users = cur.fetchall()
    ws[f'A{str(len(users))}'] = message.chat.id
    ws[f'B{str(len(users))}'] = name
    ws[f'C{str(len(users))}'] = age
    ws[f'D{str(len(users))}'] = city
    wb.save(fn)
    conn.commit()
    cur.close()
    conn.close()
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton("Овен")
    btn2 = types.KeyboardButton("Телец")
    markup.row(btn1, btn2)
    btn3 = types.KeyboardButton("Близнецы")
    btn4 = types.KeyboardButton("Рак")
    markup.row(btn3, btn4)
    btn5 = types.KeyboardButton("Лев")
    btn6 = types.KeyboardButton("Дева")
    markup.row(btn5, btn6)
    btn7 = types.KeyboardButton("Весы")
    btn8 = types.KeyboardButton("Скорпион")
    markup.row(btn7, btn8)
    btn9 = types.KeyboardButton("Стрелец")
    btn10 = types.KeyboardButton("Козерог")
    markup.row(btn9, btn10)
    btn11 = types.KeyboardButton("Водолей")
    btn12 = types.KeyboardButton("Рыбы")
    markup.row(btn11, btn12)
    bot.send_message(message.chat.id, "Выбери свой знак зодиака",reply_markup= markup)



@bot.message_handler(content_types='text')
def message_reply(message):
    """
    Делает проверку на знак зодиака,парсит и выводит данные
    :param message: знак зодиака
    :return: гороскоп
    """
    if message.text == "Овен":
        url = 'https://horoscopes.rambler.ru/aries/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Телец":
        url = 'https://horoscopes.rambler.ru/taurus/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Близнецы":
        url = 'https://horoscopes.rambler.ru/gemini/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Рак":
        url = 'https://horoscopes.rambler.ru/cancer/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Лев":
        url = 'https://horoscopes.rambler.ru/leo/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Дева":
        url = 'https://horoscopes.rambler.ru/virgo/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Весы":
        url = 'https://horoscopes.rambler.ru/libra/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Скорпион":
        url = 'https://horoscopes.rambler.ru/scorpio/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Стрелец":
        url = 'https://horoscopes.rambler.ru/sagittarius/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Козерог":
        url = 'https://horoscopes.rambler.ru/capricorn/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Водолей":
        url = 'https://horoscopes.rambler.ru/aquarius/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)
    elif message.text == "Рыбы":
        url = 'https://horoscopes.rambler.ru/pisces/'
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        horoscope_text = soup.find('p', class_='mtZOt').text
        bot.send_message(message.chat.id, horoscope_text)



bot.infinity_polling()